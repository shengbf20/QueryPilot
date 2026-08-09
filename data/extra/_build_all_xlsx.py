"""Merge easy/medium/hard into data/extra/Q&A_all.xlsx (Step 2f)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from querypilot.eval.dataset import load_qa_cases, load_qa_cases_many

DIR = Path(__file__).resolve().parent
PARTS = (
    DIR / "Q&A_easy.xlsx",
    DIR / "Q&A_medium.xlsx",
    DIR / "Q&A_hard.xlsx",
)
OUT = DIR / "Q&A_all.xlsx"
HEADERS = ["序号", "问题", "SQL", "难度", "theme"]


def main() -> None:
    for p in PARTS:
        if not p.exists():
            raise SystemExit(f"missing part: {p}")

    many = load_qa_cases_many(PARTS)
    assert len(many) == 36, len(many)
    ids = [c.id for c in many]
    assert ids[:10] == [f"E{i:02d}" for i in range(1, 11)]
    assert ids[10:24] == [f"M{i:02d}" for i in range(1, 15)]
    assert ids[24:] == [f"H{i:02d}" for i in range(1, 13)]
    assert len(set(ids)) == 36

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for path in PARTS:
        src = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = src.active.iter_rows(values_only=True)
            header = next(rows)
            # expect same logical columns
            for row in rows:
                if row is None or not any(c is not None for c in row):
                    continue
                ws.append(list(row[:5]))
        finally:
            src.close()

    wb.save(OUT)
    print(f"wrote {OUT}")

    all_cases = load_qa_cases(OUT)
    assert len(all_cases) == 36, len(all_cases)
    assert [c.id for c in all_cases] == ids
    by_diff = {}
    for c in all_cases:
        by_diff[c.difficulty or "?"] = by_diff.get(c.difficulty or "?", 0) + 1
    assert by_diff == {"简单": 10, "中等": 14, "困难": 12}, by_diff
    assert all(c.extras.get("theme") for c in all_cases)
    assert all(c.question and c.gold_sql for c in all_cases)

    # many vs all must match
    for a, b in zip(many, all_cases, strict=True):
        assert a.id == b.id
        assert a.question == b.question
        assert a.gold_sql == b.gold_sql
        assert a.difficulty == b.difficulty
        assert a.extras.get("theme") == b.extras.get("theme")

    print("OK total=36", by_diff)
    print("ids", ids)


if __name__ == "__main__":
    main()
