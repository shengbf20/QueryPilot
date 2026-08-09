"""Tests for gold Q&A loading (phase-3 step 1)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from querypilot.config import get_settings
from querypilot.eval import (
    EvalCase,
    cases_from_records,
    default_qa_path,
    load_qa_cases,
    load_qa_cases_many,
)


def test_default_qa_path_points_under_data():
    path = default_qa_path()
    assert path == get_settings().data_dir / "Q&A.xlsx"


def test_cases_from_records_canonical_and_chinese_keys():
    cases = cases_from_records(
        [
            {"id": "a", "question": "有多少女性？", "gold_sql": "SELECT 1", "difficulty": "easy"},
            {"序号": 2, "问题": "资产分布", "SQL": "SELECT 2"},
            {"question": "缺 SQL", "gold_sql": ""},  # skipped
            {"question": "", "gold_sql": "SELECT 3"},  # skipped
        ]
    )
    assert len(cases) == 2
    assert cases[0] == EvalCase(
        id="a",
        question="有多少女性？",
        gold_sql="SELECT 1",
        difficulty="easy",
    )
    assert cases[1].id == "2"
    assert cases[1].question == "资产分布"
    assert cases[1].gold_sql == "SELECT 2"
    assert cases[1].difficulty is None


def test_cases_from_records_auto_id():
    cases = cases_from_records([{"question": "q", "gold_sql": "SELECT 1"}])
    assert cases[0].id == "1"


def test_cases_from_records_accepts_extended_aliases():
    """P0: records path must honor the same header aliases as xlsx loading."""
    cases = cases_from_records(
        [
            {"问句": "q1", "标准答案": "SELECT 1", "复杂度": "hard"},
            {"自然语言": "q2", "answer_sql": "SELECT 2", "case_id": "c2"},
            {"query": "q3", "标准sql": "SELECT 3", "level": "easy"},
        ]
    )
    assert len(cases) == 3
    assert cases[0] == EvalCase(
        id="1",
        question="q1",
        gold_sql="SELECT 1",
        difficulty="hard",
    )
    assert cases[1].id == "c2"
    assert cases[1].question == "q2"
    assert cases[1].gold_sql == "SELECT 2"
    assert cases[2].question == "q3"
    assert cases[2].gold_sql == "SELECT 3"
    assert cases[2].difficulty == "easy"


def test_cases_from_records_keeps_extras():
    """P1: non-canonical columns are preserved for downstream review/tags."""
    cases = cases_from_records(
        [{"question": "q", "gold_sql": "SELECT 1", "note": "x", "tag": 9}]
    )
    assert cases[0].extras == {"note": "x", "tag": 9}


def test_load_qa_cases_missing_file(tmp_path: Path):
    missing = tmp_path / "missing.xlsx"
    with pytest.raises(FileNotFoundError, match="Gold Q&A file not found"):
        load_qa_cases(missing)


def test_load_qa_cases_unsupported_suffix(tmp_path: Path):
    bad = tmp_path / "qa.csv"
    bad.write_text("question,sql\nq,SELECT 1\n", encoding="utf-8")
    with pytest.raises(ValueError, match="Unsupported gold Q&A format"):
        load_qa_cases(bad)


def test_load_qa_cases_requires_question_and_sql(tmp_path: Path):
    path = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "note"])
    ws.append([1, "x"])
    wb.save(path)
    with pytest.raises(ValueError, match="question \\+ SQL"):
        load_qa_cases(path)


def test_load_qa_cases_from_temp_xlsx(tmp_path: Path):
    path = tmp_path / "qa.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["序号", "问题", "SQL", "难度", "theme"])
    ws.append([1, "女性客户数", "SELECT COUNT(*) FROM t", "simple", "gender_count"])
    ws.append([2, "资产汇总", "SELECT SUM(x) FROM t", None, None])
    wb.save(path)

    cases = load_qa_cases(path)
    assert len(cases) == 2
    assert cases[0].id == "1"
    assert cases[0].question == "女性客户数"
    assert "COUNT(*)" in cases[0].gold_sql
    assert cases[0].difficulty == "simple"
    assert cases[0].extras.get("theme") == "gender_count"
    assert cases[1].gold_sql.startswith("SELECT SUM")


def test_load_qa_cases_many_merges_in_order(tmp_path: Path):
    def _write(name: str, rows: list[list]) -> Path:
        path = tmp_path / name
        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "问题", "SQL", "难度", "theme"])
        for row in rows:
            ws.append(row)
        wb.save(path)
        return path

    easy = _write(
        "easy.xlsx",
        [["E01", "q1", "SELECT 1", "简单", "status_filter"]],
    )
    medium = _write(
        "medium.xlsx",
        [["M01", "q2", "SELECT 2", "中等", "credit_hold"]],
    )
    cases = load_qa_cases_many([easy, medium])
    assert [c.id for c in cases] == ["E01", "M01"]
    assert cases[0].difficulty == "简单"
    assert cases[0].extras.get("theme") == "status_filter"
    assert cases[1].extras.get("theme") == "credit_hold"


def test_load_qa_cases_many_requires_paths():
    with pytest.raises(ValueError, match="at least one path"):
        load_qa_cases_many([])


def test_load_qa_cases_extended_aliases_xlsx(tmp_path: Path):
    """P1: xlsx header aliases beyond 问题/SQL must map to EvalCase fields."""
    path = tmp_path / "qa_aliases.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["case_id", "自然语言", "标准答案", "复杂度"])
    ws.append(["c1", "女性客户数", "SELECT COUNT(*) FROM t", "hard"])
    wb.save(path)

    cases = load_qa_cases(path)
    assert len(cases) == 1
    assert cases[0].id == "c1"
    assert cases[0].question == "女性客户数"
    assert cases[0].gold_sql == "SELECT COUNT(*) FROM t"
    assert cases[0].difficulty == "hard"


def test_load_qa_cases_empty_workbook(tmp_path: Path):
    """P1: empty sheet yields no cases (not an exception)."""
    path = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.save(path)
    assert load_qa_cases(path) == []


def test_load_competition_qa_xlsx():
    path = default_qa_path()
    if not path.exists():
        pytest.skip("data/Q&A.xlsx not present")

    cases = load_qa_cases()
    assert len(cases) >= 1
    assert all(c.question and c.gold_sql for c in cases)
    # Competition workbook currently has 7 labeled items.
    assert len(cases) == 7
    assert cases[0].id == "1"
    assert "客户" in cases[0].question or "学历" in cases[0].question
    assert "select" in cases[0].gold_sql.lower()
    assert all("select" in c.gold_sql.lower() or "with" in c.gold_sql.lower() for c in cases)
    assert all(c.difficulty is None for c in cases)
