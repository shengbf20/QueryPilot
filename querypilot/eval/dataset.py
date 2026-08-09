"""Load gold Q&A pairs into standardized EvalCase list."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from querypilot.config import get_settings
from querypilot.eval.models import EvalCase

# Header aliases → canonical field (casefold match on stripped header text).
_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "id": ("序号", "id", "case_id", "qid", "no", "#"),
    "question": ("问题", "question", "query", "问句", "自然语言"),
    "gold_sql": ("sql", "gold_sql", "标准sql", "答案sql", "标准答案", "answer_sql"),
    "difficulty": ("难度", "difficulty", "level", "复杂度"),
}


def default_qa_path() -> Path:
    return get_settings().data_dir / "Q&A.xlsx"


def _norm_header(value: Any) -> str:
    return str(value or "").strip().casefold()


def _alias_lookup() -> dict[str, str]:
    """Normalized header alias → canonical field name."""
    lookup: dict[str, str] = {}
    for field, aliases in _FIELD_ALIASES.items():
        for alias in aliases:
            lookup[_norm_header(alias)] = field
    return lookup


_ALIAS_LOOKUP = _alias_lookup()


def _build_header_map(headers: Sequence[Any]) -> dict[str, int]:
    """Map canonical field name → column index."""
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        field = _ALIAS_LOOKUP.get(_norm_header(header))
        if field is not None and field not in mapping:
            mapping[field] = idx
    return mapping


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def cases_from_records(records: Iterable[Mapping[str, Any]]) -> list[EvalCase]:
    """Build EvalCase list from dict-like records (tests / alternate formats)."""
    cases: list[EvalCase] = []
    for i, row in enumerate(records, start=1):
        canonical: dict[str, Any] = {}
        extras: dict[str, Any] = {}
        for key, value in row.items():
            field = _ALIAS_LOOKUP.get(_norm_header(key))
            if field is None:
                extras[key] = value
                continue
            if field not in canonical:  # first alias wins (same as xlsx header map)
                canonical[field] = value

        question = _cell_str(canonical.get("question"))
        gold_sql = _cell_str(canonical.get("gold_sql"))
        if not question or not gold_sql:
            continue
        case_id = _cell_str(canonical.get("id")) or str(i)
        difficulty = _cell_str(canonical.get("difficulty")) or None
        cases.append(
            EvalCase(
                id=case_id,
                question=question,
                gold_sql=gold_sql,
                difficulty=difficulty,
                extras=extras,
            )
        )
    return cases


def _rows_from_xlsx(path: Path) -> tuple[tuple[Any, ...], list[tuple[Any, ...]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency declared in pyproject
        raise ImportError(
            "Reading Q&A.xlsx requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return (), []
        data = [tuple(r) for r in rows_iter if r is not None and any(c is not None for c in r)]
        return tuple(header), data
    finally:
        wb.close()


def load_qa_cases(path: Path | str | None = None) -> list[EvalCase]:
    """Load gold Q&A cases from an Excel workbook (default: data/Q&A.xlsx)."""
    qa_path = Path(path) if path is not None else default_qa_path()
    if not qa_path.exists():
        raise FileNotFoundError(
            f"Gold Q&A file not found: {qa_path}. Place Q&A.xlsx under data/."
        )
    if qa_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported gold Q&A format: {qa_path.suffix} (expect .xlsx)")

    headers, data_rows = _rows_from_xlsx(qa_path)
    if not headers:
        return []

    colmap = _build_header_map(headers)
    if "question" not in colmap or "gold_sql" not in colmap:
        raise ValueError(
            "Q&A header must include question + SQL columns "
            f"(got: {list(headers)}; mapped: {sorted(colmap)})"
        )

    mapped_idxs = set(colmap.values())
    records: list[dict[str, Any]] = []
    for row in data_rows:
        record: dict[str, Any] = {}
        for field, idx in colmap.items():
            record[field] = row[idx] if idx < len(row) else None
        # Non-canonical columns (e.g. theme) → EvalCase.extras via cases_from_records.
        for idx, header in enumerate(headers):
            if idx in mapped_idxs:
                continue
            key = _cell_str(header) or f"col_{idx}"
            if idx < len(row):
                record[key] = row[idx]
        records.append(record)
    return cases_from_records(records)


def load_qa_cases_many(paths: Sequence[Path | str]) -> list[EvalCase]:
    """Load and concatenate gold cases from multiple workbooks (file order preserved)."""
    if not paths:
        raise ValueError("load_qa_cases_many requires at least one path")
    cases: list[EvalCase] = []
    for path in paths:
        cases.extend(load_qa_cases(path))
    return cases
